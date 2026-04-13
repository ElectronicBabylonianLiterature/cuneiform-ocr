"""
Generate tablet_side_manual_review.xlsx from the visualization images
produced by test_divide_photo.ipynb (in division_checking_photos/).

Run with:
    python generate_manual_review_excel.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side as XlSide
)
from openpyxl.utils import get_column_letter

VIS_DIR  = Path('division_checking_photos')
OUT_FILE = Path('tablet_side_manual_review.xlsx')

# ── collect files ──────────────────────────────────────────────────────────────
vis_files = sorted(
    p for p in VIS_DIR.iterdir()
    if p.suffix.lower() in ('.jpg', '.jpeg', '.png')
)
if not vis_files:
    raise FileNotFoundError(f"No images found in {VIS_DIR.resolve()}")

fragment_ids = [p.stem for p in vis_files]
print(f"Found {len(vis_files)} visualization files in {VIS_DIR}/")

# ── style helpers ──────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
SUBHDR_FILL = PatternFill('solid', fgColor='2E75B6')
SUBHDR_FONT = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL   = PatternFill('solid', fgColor='DCE6F1')
SUM_FILL   = PatternFill('solid', fgColor='FFF2CC')
SUM_FONT   = Font(bold=True, size=10)
WARN_FILL  = PatternFill('solid', fgColor='FFE0E0')
C_ALIGN    = Alignment(horizontal='center', vertical='center', wrap_text=True)
L_ALIGN    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
R_ALIGN    = Alignment(horizontal='right',  vertical='center')

def thin_border():
    s = XlSide(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def apply(cell, *, font=None, fill=None, align=None, number_format=None):
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if align:         cell.alignment     = align
    if number_format: cell.number_format = number_format
    cell.border = thin_border()

# ── create workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Manual Review ─────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Manual Review'

HEADERS = [
    '#',
    'Fragment ID',
    'Visualization Path',
    'Predicted\nSides',
    'True Total\nSides',
    'Missed\nDetections (FN)',
    'False\nDetections (FP)',
    'Notes',
]
COL_WIDTHS = [5, 22, 62, 10, 12, 16, 16, 35]
# Column letters for formula references
C_PRED   = 'D'
C_TRUE   = 'E'
C_FN     = 'F'
C_FP     = 'G'

# Header row
for ci, (hdr, wid) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=ci, value=hdr)
    apply(cell, font=HDR_FONT, fill=HDR_FILL, align=C_ALIGN)
    ws.column_dimensions[get_column_letter(ci)].width = wid
ws.row_dimensions[1].height = 36

# Data rows
first_data_row = 2
last_data_row  = 1 + len(vis_files)

for ri, (fid, vis_path) in enumerate(zip(fragment_ids, vis_files), start=2):
    row_data = [
        ri - 1,
        fid,
        str(vis_path.resolve()),
        '',  # Predicted Sides  — count from visualization
        '',  # True Total Sides — manual
        '',  # FN
        '',  # FP
        '',  # Notes
    ]
    use_alt = (ri % 2 == 0)
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        alt  = ALT_FILL if use_alt else PatternFill()
        apply(cell, fill=alt, align=L_ALIGN if ci in (3, 8) else C_ALIGN)
        # Numeric hint for manual-fill columns
        if ci in (4, 5, 6, 7) and val == '':
            cell.number_format = '0'
    ws.row_dimensions[ri].height = 18

# ── Summary block ──────────────────────────────────────────────────────────────
SUM_ROW   = last_data_row + 2   # blank gap
STATS_ROW = SUM_ROW + 1         # Precision / Recall

def rng(col):
    return f"{col}{first_data_row}:{col}{last_data_row}"

sum_labels = [
    ('Total Predicted Sides',     C_PRED,  f'=SUM({rng(C_PRED)})'),
    ('Total True Sides',          C_TRUE,  f'=SUM({rng(C_TRUE)})'),
    ('Total Missed (FN)',         C_FN,    f'=SUM({rng(C_FN)})'),
    ('Total False Detections (FP)', C_FP,  f'=SUM({rng(C_FP)})'),
]

# Row header
cell = ws.cell(row=SUM_ROW, column=1, value='Summary')
apply(cell, font=SUM_FONT, fill=SUM_FILL, align=C_ALIGN)
ws.merge_cells(f'A{SUM_ROW}:C{SUM_ROW}')
ws.row_dimensions[SUM_ROW].height = 20

for label, col, formula in sum_labels:
    col_idx = ord(col) - ord('A') + 1
    # label
    lbl_cell = ws.cell(row=SUM_ROW, column=col_idx, value=label)
    apply(lbl_cell, font=SUM_FONT, fill=SUM_FILL, align=C_ALIGN)
    # formula (one row below label)
    val_cell = ws.cell(row=STATS_ROW, column=col_idx, value=formula)
    apply(val_cell, font=SUM_FONT, fill=SUM_FILL, align=C_ALIGN)

# Recall and Precision derived formulas (columns F and G of summary)
pred_cell  = f'{C_PRED}{STATS_ROW}'
true_cell  = f'{C_TRUE}{STATS_ROW}'
fn_cell    = f'{C_FN}{STATS_ROW}'
fp_cell    = f'{C_FP}{STATS_ROW}'

recall_col    = ord(C_FN) - ord('A') + 1   # same column as FN
precision_col = ord(C_FP) - ord('A') + 1   # same column as FP

RECALL_ROW    = STATS_ROW + 1
PRECISION_ROW = STATS_ROW + 2

for row_i, label, formula, note in [
    (RECALL_ROW,    'Recall',
     f'=IF({true_cell}>0, 1-{fn_cell}/{true_cell}, "")',
     '= 1 − FN / True Total'),
    (PRECISION_ROW, 'Precision',
     f'=IF({pred_cell}>0, 1-{fp_cell}/{pred_cell}, "")',
     '= 1 − FP / Predicted'),
]:
    # label spans A-C
    lbl = ws.cell(row=row_i, column=1, value=label)
    apply(lbl, font=SUM_FONT, fill=SUM_FILL, align=C_ALIGN)
    ws.merge_cells(f'A{row_i}:C{row_i}')

    val = ws.cell(row=row_i, column=recall_col if label == 'Recall' else precision_col,
                  value=formula)
    apply(val, font=Font(bold=True, size=11), fill=SUM_FILL, align=C_ALIGN,
          number_format='0.0%')

    note_cell = ws.cell(row=row_i, column=8, value=note)
    apply(note_cell, fill=SUM_FILL, align=L_ALIGN)
    ws.row_dimensions[row_i].height = 20

ws.row_dimensions[STATS_ROW].height = 20

ws.freeze_panes = 'A2'

# ── Sheet 2: Instructions ──────────────────────────────────────────────────────
ws2 = wb.create_sheet('Instructions')
ws2.column_dimensions['A'].width = 80
lines = [
    ('Tablet-Side Extraction — Manual Review Instructions', True),
    ('', False),
    ('Purpose', True),
    ('  Validate divide_tablet_photo by visual inspection, without relying on GT annotations.', False),
    ('  GT annotations are known to be incomplete (only some sides are labeled per image),', False),
    ('  so this manual review serves as the primary quantitative evaluation.', False),
    ('', False),
    ('How to review each row', True),
    ('  1. Open the visualization image at the path shown in column C.', False),
    ('     The image shows the photo with detected side regions numbered and boxed.', False),
    ('  2. Column D — Predicted Sides: count the numbered boxes visible in the image.', False),
    ('  3. Column E — True Total Sides: how many tablet sides are actually visible in the photo.', False),
    ('  4. Column F — Missed Detections (FN): sides visible but NOT detected.', False),
    ('  5. Column G — False Detections (FP): detected regions that are NOT real sides.', False),
    ('  6. Column H — Notes: any observations (e.g. "edge cut off", "background noise").', False),
    ('', False),
    ('Correctness criteria', True),
    ('  A detection is CORRECT if it visually isolates one tablet side as its own crop.', False),
    ('  Judge by appearance — no IoU threshold is applied.', False),
    ('  A result is WRONG if:', False),
    ('    - A visible side is not detected at all  (→ Missed / FN)', False),
    ('    - A detected crop does not correspond to a real side  (→ False / FP)', False),
    ('    - Two sides are merged into one crop  (counts as 1 FN + 0 FP)', False),
    ('    - One side is split into multiple crops  (counts as 0 FN + N-1 FP)', False),
    ('', False),
    ('Summary statistics', True),
    ('  The yellow summary block at the bottom of the sheet computes:', False),
    ('    Recall    = 1 − Total_FN / Total_True_Sides', False),
    ('    Precision = 1 − Total_FP / Total_Predicted_Sides', False),
    ('  These update automatically as you fill in columns D–G.', False),
]
l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
for i, (text, bold) in enumerate(lines, 1):
    cell = ws2.cell(row=i, column=1, value=text)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = l_align

wb.save(OUT_FILE)
print(f"Saved: {OUT_FILE.resolve()}")
print(f"  Rows: {len(vis_files)} samples + summary block")
